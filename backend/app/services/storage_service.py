from __future__ import annotations

import base64
import hashlib
import mimetypes
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from fastapi import UploadFile
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.relatorio import RelatorioArquivo

settings = get_settings()

ALLOWED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".txt", ".docx", ".xlsx", ".xls"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
TEXT_EXTENSIONS = {".txt"}


@dataclass
class StoredFileContext:
    arquivo_id: int
    filename: str
    extension: str
    path: Path
    mime_type: str | None
    size_bytes: int
    sha256: str
    extracted_text: str = ""
    data_url: str | None = None


def _safe_filename(filename: str) -> str:
    keep = []
    for ch in filename:
        if ch.isalnum() or ch in {".", "-", "_", " ", "(", ")"}:
            keep.append(ch)
        else:
            keep.append("_")
    return "".join(keep).strip() or "arquivo"


def _extract_txt(path: Path) -> str:
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        pages: list[str] = []
        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"\n--- Página {i} ---\n{text}")
        return "\n".join(pages).strip()
    except Exception as exc:  # noqa: BLE001
        return f"[Não foi possível extrair texto do PDF: {exc}]"


def _extract_docx(path: Path) -> str:
    try:
        from docx import Document

        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as exc:  # noqa: BLE001
        return f"[Não foi possível extrair texto do DOCX: {exc}]"


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True, read_only=True)
        parts: list[str] = []
        for ws in wb.worksheets:
            parts.append(f"\n--- Aba: {ws.title} ---")
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() if v is not None else "" for v in row]
                if any(values):
                    parts.append(" | ".join(values))
        wb.close()
        return "\n".join(parts).strip()
    except Exception as exc:  # noqa: BLE001
        return f"[Não foi possível extrair texto do Excel: {exc}]"


def extract_text_from_file(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".txt":
        return _extract_txt(path)
    if ext == ".pdf":
        return _extract_pdf(path)
    if ext == ".docx":
        return _extract_docx(path)
    if ext in {".xlsx", ".xls"}:
        return _extract_xlsx(path)
    return ""


def image_data_url(path: Path, mime_type: str | None = None) -> str | None:
    if path.suffix.lower() not in IMAGE_EXTENSIONS:
        return None
    mime = mime_type or mimetypes.guess_type(path.name)[0] or "image/png"
    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{data}"


async def save_uploads(
    db: Session,
    relatorio_id: int,
    arquivos: Iterable[UploadFile] | None,
) -> list[StoredFileContext]:
    contexts: list[StoredFileContext] = []
    target_dir = settings.upload_path / str(relatorio_id)
    target_dir.mkdir(parents=True, exist_ok=True)

    for upload in arquivos or []:
        original_name = upload.filename or "arquivo"
        extension = Path(original_name).suffix.lower()
        if extension and extension not in ALLOWED_EXTENSIONS:
            raise ValueError(f"Extensão não permitida: {extension}")

        content = await upload.read()
        sha256 = hashlib.sha256(content).hexdigest()
        filename = _safe_filename(original_name)
        path = target_dir / filename

        if path.exists():
            stem = path.stem
            suffix = path.suffix
            path = target_dir / f"{stem}_{sha256[:8]}{suffix}"

        path.write_bytes(content)
        mime_type = upload.content_type or mimetypes.guess_type(filename)[0]

        model = RelatorioArquivo(
            relatorio_id=relatorio_id,
            nome_arquivo=original_name,
            tipo_arquivo=mime_type,
            extensao=extension,
            storage_path=str(path),
            tamanho_bytes=len(content),
            hash_arquivo=sha256,
        )
        db.add(model)
        db.flush()

        extracted_text = extract_text_from_file(path)
        data_url = image_data_url(path, mime_type)
        contexts.append(
            StoredFileContext(
                arquivo_id=model.id,
                filename=original_name,
                extension=extension,
                path=path,
                mime_type=mime_type,
                size_bytes=len(content),
                sha256=sha256,
                extracted_text=extracted_text,
                data_url=data_url,
            )
        )

    db.flush()
    return contexts


def build_files_context_text(files: list[StoredFileContext]) -> str:
    if not files:
        return "Nenhum arquivo anexado."

    parts: list[str] = []
    for item in files:
        parts.append(
            f"Arquivo: {item.filename}\n"
            f"Extensão: {item.extension}\n"
            f"Tipo: {item.mime_type or 'não informado'}\n"
            f"Hash: {item.sha256}\n"
            f"Texto extraído:\n{item.extracted_text or '[sem texto extraído automaticamente]'}"
        )
    return "\n\n".join(parts)
